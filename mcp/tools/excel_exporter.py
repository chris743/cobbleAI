"""Generate downloadable Excel files from query results."""

import uuid
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .query_executor import QueryExecutor


class ExcelExporter:
    """Generate downloadable Excel files from query results."""

    def __init__(self, export_path: str = "./exports"):
        self.export_path = Path(export_path)
        self.export_path.mkdir(exist_ok=True)

    def _write_sheet(self, ws, columns: list[str], rows: list[list]):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(str(col_name))
            for row in rows[:100]:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1] or "")))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 50)

    def _make_filename(self, filename: str) -> str:
        file_id = str(uuid.uuid4())[:8]
        if not filename:
            filename = f"export_{file_id}"
        safe_name = "".join(c for c in filename if c.isalnum() or c in "_- ").strip()
        if not safe_name:
            safe_name = f"export_{file_id}"
        return f"{safe_name}_{file_id}.xlsx"

    def export(self, columns: list[str] = None, rows: list[list] = None,
               filename: str = "", sheets: list[dict] = None) -> dict:
        if sheets:
            sheet_list = sheets
        elif columns and rows:
            sheet_list = [{"name": "Data", "columns": columns, "rows": rows}]
        else:
            return {"success": False, "error": "No data to export"}

        for s in sheet_list:
            if not s.get("columns") or not s.get("rows"):
                return {"success": False, "error": f"Sheet '{s.get('name', '?')}' has no data"}

        xlsx_name = self._make_filename(filename)

        wb = Workbook()
        wb.remove(wb.active)

        total_rows = 0
        for sheet_def in sheet_list:
            sheet_name = str(sheet_def.get("name", "Sheet"))[:31]
            ws = wb.create_sheet(title=sheet_name)
            self._write_sheet(ws, sheet_def["columns"], sheet_def["rows"])
            total_rows += len(sheet_def["rows"])

        filepath = self.export_path / xlsx_name
        wb.save(filepath)

        return {
            "success": True,
            "file_id": xlsx_name,
            "download_url": f"/download/{xlsx_name}",
            "row_count": total_rows,
            "sheet_count": len(sheet_list),
            "message": f"Excel file ready: {xlsx_name} ({len(sheet_list)} sheet{'s' if len(sheet_list) != 1 else ''})"
        }

    def export_from_queries(self, queries: list[dict], filename: str = "",
                            split_by_column: str = None) -> dict:
        if not queries:
            return {"success": False, "error": "No queries provided"}

        executor = QueryExecutor()
        xlsx_name = self._make_filename(filename)

        wb = Workbook()
        wb.remove(wb.active)

        total_rows = 0
        sheet_results = []

        if split_by_column:
            q = queries[0]
            sql = q.get("sql", "")
            database = q.get("database")
            if not sql:
                return {"success": False, "error": "No SQL provided"}

            result = executor.execute(sql, database=database, _internal_limit=50000)
            if not result.get("success"):
                return {"success": False, "error": result.get("error", "Query failed")}

            columns = result["columns"]
            rows = result["rows"]
            if not rows:
                return {"success": False, "error": "Query returned no data"}

            split_col = split_by_column.strip()
            try:
                split_idx = next(i for i, c in enumerate(columns) if c.lower() == split_col.lower())
            except StopIteration:
                return {"success": False, "error": f"Column '{split_by_column}' not found. Available: {columns}"}

            out_columns = [c for i, c in enumerate(columns) if i != split_idx]

            groups = OrderedDict()
            for row in rows:
                key = str(row[split_idx] or "Other")
                if key not in groups:
                    groups[key] = []
                groups[key].append([v for i, v in enumerate(row) if i != split_idx])

            for group_name, group_rows in groups.items():
                sheet_name = str(group_name)[:31]
                ws = wb.create_sheet(title=sheet_name)
                self._write_sheet(ws, out_columns, group_rows)
                total_rows += len(group_rows)
                sheet_results.append({"sheet": sheet_name, "rows": len(group_rows)})
        else:
            for q in queries:
                sheet_name = str(q.get("name", "Sheet"))[:31]
                sql = q.get("sql", "")
                database = q.get("database")

                if not sql:
                    sheet_results.append({"sheet": sheet_name, "error": "No SQL provided"})
                    continue

                result = executor.execute(sql, database=database, _internal_limit=50000)
                if not result.get("success"):
                    sheet_results.append({"sheet": sheet_name, "error": result.get("error", "Query failed")})
                    continue

                columns = result["columns"]
                rows = result["rows"]
                if not rows:
                    sheet_results.append({"sheet": sheet_name, "rows": 0, "note": "No data returned"})
                    continue

                ws = wb.create_sheet(title=sheet_name)
                self._write_sheet(ws, columns, rows)
                total_rows += len(rows)
                sheet_results.append({"sheet": sheet_name, "rows": len(rows)})

        if total_rows == 0:
            return {"success": False, "error": "All queries returned no data", "details": sheet_results}

        filepath = self.export_path / xlsx_name
        wb.save(filepath)

        return {
            "success": True,
            "file_id": xlsx_name,
            "download_url": f"/download/{xlsx_name}",
            "row_count": total_rows,
            "sheet_count": len([s for s in sheet_results if s.get("rows", 0) > 0]),
            "sheets": sheet_results,
            "message": f"Excel file ready: {xlsx_name} ({total_rows} total rows)"
        }


TOOL_DEFINITIONS = [
    {
        "name": "export_excel",
        "description": "Export data to a downloadable Excel (.xlsx) file. Supports single-sheet (columns/rows) or multi-sheet (sheets array) exports. For multi-sheet, pass a 'sheets' array where each element has name, columns, and rows. Always use multi-sheet when the user asks for multiple tables or breakdowns in one file. Returns a download link to include in your response.",
        "parameters": {
            "type": "object",
            "properties": {
                "columns": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Column headers (single-sheet mode). Omit if using 'sheets'."
                },
                "rows": {
                    "type": "array",
                    "items": {"type": "array"},
                    "description": "Data rows (single-sheet mode). Omit if using 'sheets'."
                },
                "sheets": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "name": {"type": "string", "description": "Sheet/tab name (max 31 chars)"},
                            "columns": {"type": "array", "items": {"type": "string"}},
                            "rows": {"type": "array", "items": {"type": "array"}}
                        },
                        "required": ["name", "columns", "rows"]
                    },
                    "description": "Array of sheets for multi-sheet export. Each sheet has a name, columns, and rows."
                },
                "filename": {
                    "type": "string",
                    "description": "Descriptive filename (without extension), e.g. 'navel_inventory_march'"
                }
            }
        }
    },
    {
        "name": "export_sql_to_excel",
        "description": "Run SQL queries and export results DIRECTLY to Excel — data goes straight from database to spreadsheet. Use this instead of export_excel for complete exports. Supports two modes: (1) Multiple queries, each becoming a sheet. (2) Single query with split_by_column to auto-split results into one sheet per distinct value of that column (e.g., split by Style to get one tab per style). Use split_by_column for production schedules.",
        "parameters": {
            "type": "object",
            "properties": {
                "queries": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "name": {"type": "string", "description": "Sheet/tab name (max 31 chars). Ignored when split_by_column is used."},
                            "sql": {"type": "string", "description": "SELECT query to run"},
                            "database": {"type": "string", "description": "Database to query (DM03 or DM01). Defaults to DM03."}
                        },
                        "required": ["name", "sql"]
                    },
                    "description": "Array of queries. When split_by_column is set, only the first query is used."
                },
                "split_by_column": {
                    "type": "string",
                    "description": "Column name to split results by. Each distinct value becomes a separate sheet tab. The column is removed from the sheet data. E.g., 'Style' splits orders into one sheet per style. For production schedules, include a Style column in the query and set this to 'Style'."
                },
                "filename": {
                    "type": "string",
                    "description": "Descriptive filename (without extension)"
                }
            },
            "required": ["queries"]
        }
    },
]


def register_handlers(exporter: ExcelExporter) -> dict:
    return {
        "export_excel": lambda p: exporter.export(
            p.get("columns"), p.get("rows"), p.get("filename", ""), p.get("sheets")
        ),
        "export_sql_to_excel": lambda p: exporter.export_from_queries(
            p.get("queries", []), p.get("filename", ""), p.get("split_by_column")
        ),
    }
