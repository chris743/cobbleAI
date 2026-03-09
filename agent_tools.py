"""
DM03 Data Warehouse Agent Tools
================================
Tool framework for AI agents to query and learn from the data warehouse.

Components:
1. Query Executor - Run validated SQL against the warehouse
2. Context Loader - Load schema documentation into agent context
3. Learning Tools - Record discoveries, corrections, and successful patterns

Usage:
    from agent_tools import AgentToolkit
    
    toolkit = AgentToolkit()
    
    # Get tools for LLM function calling
    tools = toolkit.get_tool_definitions()
    
    # Handle a tool call from the agent
    result = toolkit.handle_tool_call("execute_sql", {"sql": "SELECT ..."})
"""

import pyodbc
import yaml
import json
import os
import uuid
import httpx
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, Any
from decimal import Decimal
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Load environment variables from .env file
load_dotenv()

import living_docs as _living_docs
import customer_specs as _customer_specs

# =============================================================================
# CONFIGURATION (from .env)
# =============================================================================

CONFIG = {
    "server": os.getenv("DB_SERVER", "RDGW-CF"),
    "database": os.getenv("DB_DATABASE", "DM03"),
    "username": os.getenv("DB_USERNAME"),
    "password": os.getenv("DB_PASSWORD"),
    "trusted_connection": os.getenv("DB_TRUSTED_CONNECTION", "yes").lower() == "yes",
    "context_path": os.getenv("CONTEXT_PATH", "./data-catalog"),
    "learning_path": os.getenv("LEARNING_PATH", "./agent-learning"),
    "max_rows": int(os.getenv("MAX_ROWS", "5000")),
    "query_timeout": int(os.getenv("QUERY_TIMEOUT", "30")),
    # Harvest Planner API
    "hp_base_url": os.getenv("HP_BASE_URL", "").rstrip("/"),
    "hp_username": os.getenv("HP_USERNAME", ""),
    "hp_password": os.getenv("HP_PASSWORD", ""),
}

# Build connection strings
def _build_conn_string(database: str) -> str:
    if CONFIG["trusted_connection"]:
        return (
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={CONFIG['server']};"
            f"DATABASE={database};"
            f"Trusted_Connection=yes;"
        )
    return (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={CONFIG['server']};"
        f"DATABASE={database};"
        f"UID={CONFIG['username']};"
        f"PWD={CONFIG['password']};"
        f"Encrypt=no;"
        f"TrustServerCertificate=yes;"
        f"Application Name=DM03_Agent;"
    )

CONNECTION_STRINGS = {
    "DM03": _build_conn_string("DM03"),
    "DM01": _build_conn_string("DM01"),
}
CONNECTION_STRING = CONNECTION_STRINGS["DM03"]  # default


# =============================================================================
# QUERY EXECUTOR
# =============================================================================

class QueryExecutor:
    """Safe, read-only SQL executor for agent queries."""
    
    FORBIDDEN_PATTERNS = [
        'INSERT', 'UPDATE', 'DELETE', 'DROP', 'TRUNCATE', 'ALTER', 'CREATE',
        'EXEC', 'EXECUTE', 'GRANT', 'REVOKE', 'INTO', '--', '/*', 'xp_', 'sp_'
    ]
    
    def __init__(self, connection_string: str = CONNECTION_STRING):
        self.connection_string = connection_string
        self.max_rows = CONFIG["max_rows"]
        self.timeout = CONFIG["query_timeout"]
    
    def _validate_query(self, sql: str) -> tuple[bool, str]:
        sql_upper = sql.upper().strip()
        
        if not (sql_upper.startswith('SELECT') or sql_upper.startswith('WITH')):
            return False, "Query must start with SELECT or WITH"
        
        for pattern in self.FORBIDDEN_PATTERNS:
            if pattern in sql_upper:
                return False, f"Query contains forbidden keyword: {pattern}"
        
        return True, ""
    
    def _serialize(self, val: Any) -> Any:
        if val is None:
            return None
        elif isinstance(val, datetime):
            return val.isoformat()
        elif isinstance(val, Decimal):
            return float(val)
        elif isinstance(val, bytes):
            return val.hex()
        return val
    
    def execute(self, sql: str, max_rows: Optional[int] = None, database: str = None,
                _internal_limit: int = None) -> dict:
        effective_max = _internal_limit or self.max_rows
        max_rows = min(max_rows or effective_max, effective_max)

        is_valid, error = self._validate_query(sql)
        if not is_valid:
            return {"success": False, "error": error, "sql": sql}

        conn_str = self.connection_string
        if database and database.upper() in CONNECTION_STRINGS:
            conn_str = CONNECTION_STRINGS[database.upper()]

        try:
            conn = pyodbc.connect(conn_str, timeout=self.timeout)
            cursor = conn.cursor()
            cursor.execute(sql)
            
            columns = [desc[0] for desc in cursor.description]
            rows = cursor.fetchmany(max_rows + 1)
            truncated = len(rows) > max_rows
            if truncated:
                rows = rows[:max_rows]
            
            serialized = [[self._serialize(v) for v in row] for row in rows]
            conn.close()
            
            return {
                "success": True,
                "columns": columns,
                "rows": serialized,
                "row_count": len(serialized),
                "truncated": truncated,
                "sql": sql
            }
        except Exception as e:
            return {"success": False, "error": str(e), "sql": sql}


# =============================================================================
# CONTEXT LOADER
# =============================================================================

class ContextLoader:
    """Load and search schema documentation."""
    
    def __init__(self, context_path: str = CONFIG["context_path"]):
        self.context_path = Path(context_path)
    
    def load_file(self, filename: str) -> dict:
        """Load a specific YAML file from context."""
        filepath = self.context_path / filename
        if not filepath.exists():
            return {"error": f"File not found: {filename}"}
        
        with open(filepath, 'r') as f:
            return yaml.safe_load(f)
    
    def load_system_architecture(self) -> dict:
        """Load the system architecture reference."""
        return self.load_file("system_architecture.yaml")
    
    def load_glossary(self) -> dict:
        """Load the business glossary."""
        return self.load_file("glossary.yaml")

    def load_size_dictionary(self, commodity: str = None) -> dict:
        """Load size reference data, optionally filtered to one commodity."""
        data = self.load_file("size_dictionary.yaml")
        if "error" in data:
            return data
        if commodity:
            commodity_upper = commodity.upper()
            commodities = data.get("commodities", {})
            for key, val in commodities.items():
                if key.upper() == commodity_upper:
                    return {"commodity": key, **val, "special_size_codes": data.get("special_size_codes", {})}
            return {"error": f"Commodity not found: {commodity}. Available: {list(commodities.keys())}"}
        return data
    
    def load_domain(self, domain_name: str) -> dict:
        """Load a domain definition."""
        return self.load_file(f"domains/{domain_name}/_domain.yaml")
    
    def load_table_schema(self, table_name: str) -> dict:
        """Load schema documentation for a specific table."""
        # Normalize the table name for searching
        search_name = table_name.lower()
        search_name = search_name.replace("dbo.", "").replace("rpt.", "").replace("ref.", "")
        search_name = search_name.replace("vw_", "").replace("_", "")
        
        # Try to find the table file
        for yaml_file in self.context_path.rglob("*.yaml"):
            file_stem = yaml_file.stem.lower().replace("vw_", "").replace("_", "")
            
            # Check if this file matches
            if search_name in file_stem or file_stem in search_name:
                with open(yaml_file, 'r') as f:
                    content = yaml.safe_load(f)
                    # Verify it's actually a table schema
                    if isinstance(content, dict) and "columns" in content:
                        return content
        
        # Also try exact name match on the 'name' field inside files
        for yaml_file in self.context_path.rglob("*.yaml"):
            try:
                with open(yaml_file, 'r') as f:
                    content = yaml.safe_load(f)
                    if isinstance(content, dict):
                        doc_name = content.get("name", "").lower()
                        if table_name.lower() in doc_name or doc_name in table_name.lower():
                            if "columns" in content:
                                return content
            except:
                pass
        
        return {"error": f"Schema not found for table: {table_name}"}
    
    def list_tables(self) -> list[str]:
        """List all documented tables."""
        tables = []
        for yaml_file in self.context_path.rglob("*.yaml"):
            content = self.load_file(str(yaml_file.relative_to(self.context_path)))
            if isinstance(content, dict) and content.get("name", "").startswith("dbo."):
                tables.append(content["name"])
        return tables
    
    def search_context(self, query: str) -> list[dict]:
        """Search across all documentation for relevant info."""
        results = []
        query_lower = query.lower()
        
        for yaml_file in self.context_path.rglob("*.yaml"):
            try:
                with open(yaml_file, 'r') as f:
                    content = f.read()
                    if query_lower in content.lower():
                        results.append({
                            "file": str(yaml_file.relative_to(self.context_path)),
                            "preview": content[:500] + "..." if len(content) > 500 else content
                        })
            except:
                pass
        
        return results
    
    def get_full_context(self) -> str:
        """Generate complete context document for agent."""
        context_parts = []
        
        # System architecture first
        arch = self.load_system_architecture()
        if "error" not in arch:
            context_parts.append("# SYSTEM ARCHITECTURE\n" + yaml.dump(arch))
        
        # Glossary
        glossary = self.load_glossary()
        if "error" not in glossary:
            context_parts.append("# GLOSSARY\n" + yaml.dump(glossary))
        
        # All table schemas
        for yaml_file in self.context_path.rglob("*.yaml"):
            if yaml_file.name.startswith("vw_") or yaml_file.name.startswith("bininv"):
                with open(yaml_file, 'r') as f:
                    context_parts.append(f"# TABLE: {yaml_file.stem}\n" + f.read())
        
        return "\n\n---\n\n".join(context_parts)


# =============================================================================
# LEARNING TOOLS
# =============================================================================

class LearningManager:
    """Manage agent learning - queries, corrections, discoveries."""
    
    def __init__(self, learning_path: str = CONFIG["learning_path"]):
        self.learning_path = Path(learning_path)
        self.learning_path.mkdir(exist_ok=True)
        
        self.queries_file = self.learning_path / "learned_queries.yaml"
        self.corrections_file = self.learning_path / "corrections.yaml"
        self.discoveries_file = self.learning_path / "discoveries.yaml"

        # Initialize files if they don't exist
        for f in [self.queries_file, self.corrections_file, self.discoveries_file]:
            if not f.exists():
                with open(f, 'w') as file:
                    yaml.dump({"entries": []}, file)
    
    def _load_file(self, filepath: Path) -> dict:
        with open(filepath, 'r') as f:
            return yaml.safe_load(f) or {"entries": []}
    
    def _save_file(self, filepath: Path, data: dict):
        with open(filepath, 'w') as f:
            yaml.dump(data, f, default_flow_style=False, sort_keys=False)
    
    def remember_query(self, question: str, sql: str, notes: str = "") -> dict:
        """Store a successful query pattern for future reference."""
        data = self._load_file(self.queries_file)
        
        entry = {
            "timestamp": datetime.now().isoformat(),
            "question": question,
            "sql": sql,
            "notes": notes,
            "verified": False
        }
        
        data["entries"].append(entry)
        self._save_file(self.queries_file, data)
        
        return {"success": True, "message": "Query pattern saved"}
    
    def log_correction(self, table: str, column: str, 
                       wrong_assumption: str, correct_meaning: str) -> dict:
        """Log when the agent was corrected about something."""
        data = self._load_file(self.corrections_file)
        
        entry = {
            "timestamp": datetime.now().isoformat(),
            "table": table,
            "column": column,
            "wrong_assumption": wrong_assumption,
            "correct_meaning": correct_meaning,
            "applied_to_docs": False
        }
        
        data["entries"].append(entry)
        self._save_file(self.corrections_file, data)
        
        return {"success": True, "message": "Correction logged"}
    
    def record_discovery(self, category: str, key: str, value: str) -> dict:
        """Record a new discovery about the data or schema."""
        data = self._load_file(self.discoveries_file)
        
        entry = {
            "timestamp": datetime.now().isoformat(),
            "category": category,  # column_meaning, join_pattern, gotcha, business_rule
            "key": key,
            "value": value,
            "verified": False
        }
        
        data["entries"].append(entry)
        self._save_file(self.discoveries_file, data)
        
        return {"success": True, "message": "Discovery recorded"}
    
    def get_learned_queries(self, search: str = None) -> list[dict]:
        """Retrieve learned query patterns, optionally filtered."""
        data = self._load_file(self.queries_file)
        entries = data.get("entries", [])
        
        if search:
            search_lower = search.lower()
            entries = [e for e in entries 
                      if search_lower in e.get("question", "").lower() 
                      or search_lower in e.get("sql", "").lower()]
        
        return entries
    
    def get_corrections(self) -> list[dict]:
        """Get all logged corrections."""
        data = self._load_file(self.corrections_file)
        return data.get("entries", [])
    
    def get_discoveries(self, category: str = None) -> list[dict]:
        """Get recorded discoveries, optionally by category."""
        data = self._load_file(self.discoveries_file)
        entries = data.get("entries", [])

        if category:
            entries = [e for e in entries if e.get("category") == category]

        return entries



# =============================================================================
# EXCEL EXPORTER
# =============================================================================

class ExcelExporter:
    """Generate downloadable Excel files from query results."""

    def __init__(self, export_path: str = "./exports"):
        self.export_path = Path(export_path)
        self.export_path.mkdir(exist_ok=True)

    def _write_sheet(self, ws, columns: list[str], rows: list[list]):
        """Write headers and data rows to a worksheet."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths (approximate)
        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(str(col_name))
            for row in rows[:100]:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1] or "")))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 50)

    def export(self, columns: list[str] = None, rows: list[list] = None,
               filename: str = "", sheets: list[dict] = None) -> dict:
        """Export to Excel. Supports single-sheet (columns/rows) or multi-sheet (sheets array).
        Each sheet dict: { "name": "Sheet Name", "columns": [...], "rows": [[...], ...] }
        """
        # Build sheet list — support both single-sheet and multi-sheet calls
        if sheets:
            sheet_list = sheets
        elif columns and rows:
            sheet_list = [{"name": "Data", "columns": columns, "rows": rows}]
        else:
            return {"success": False, "error": "No data to export"}

        # Validate
        for s in sheet_list:
            if not s.get("columns") or not s.get("rows"):
                return {"success": False, "error": f"Sheet '{s.get('name', '?')}' has no data"}

        file_id = str(uuid.uuid4())[:8]
        if not filename:
            filename = f"export_{file_id}"
        safe_name = "".join(c for c in filename if c.isalnum() or c in "_- ").strip()
        if not safe_name:
            safe_name = f"export_{file_id}"
        xlsx_name = f"{safe_name}_{file_id}.xlsx"

        wb = Workbook()
        # Remove the default sheet created by openpyxl
        wb.remove(wb.active)

        total_rows = 0
        for sheet_def in sheet_list:
            sheet_name = str(sheet_def.get("name", "Sheet"))[:31]  # Excel max 31 chars
            ws = wb.create_sheet(title=sheet_name)
            self._write_sheet(ws, sheet_def["columns"], sheet_def["rows"])
            total_rows += len(sheet_def["rows"])

        filepath = self.export_path / xlsx_name
        wb.save(filepath)

        return {
            "success": True,
            "file_id": xlsx_name,
            "download_url": f"/download/{xlsx_name}",
            "row_count": total_rows,
            "sheet_count": len(sheet_list),
            "message": f"Excel file ready: {xlsx_name} ({len(sheet_list)} sheet{'s' if len(sheet_list) != 1 else ''})"
        }

    def export_from_queries(self, queries: list[dict], filename: str = "",
                            split_by_column: str = None) -> dict:
        """Run SQL queries directly and export results to a multi-sheet Excel file.
        Each query dict: { "name": "Sheet Name", "sql": "SELECT ...", "database": "DM03" (optional) }

        If split_by_column is set, only the FIRST query is used. Its results are split
        into separate sheets based on distinct values of that column. The column is
        removed from the sheet data (it becomes the tab name). Great for splitting
        orders by Style, EquipmentLine, Commodity, etc.
        """
        if not queries:
            return {"success": False, "error": "No queries provided"}

        executor = QueryExecutor()
        file_id = str(uuid.uuid4())[:8]
        if not filename:
            filename = f"export_{file_id}"
        safe_name = "".join(c for c in filename if c.isalnum() or c in "_- ").strip()
        if not safe_name:
            safe_name = f"export_{file_id}"
        xlsx_name = f"{safe_name}_{file_id}.xlsx"

        wb = Workbook()
        wb.remove(wb.active)

        total_rows = 0
        sheet_results = []

        if split_by_column:
            # Split mode: run first query, split results by column value
            q = queries[0]
            sql = q.get("sql", "")
            database = q.get("database")
            if not sql:
                return {"success": False, "error": "No SQL provided"}

            result = executor.execute(sql, database=database, _internal_limit=50000)
            if not result.get("success"):
                return {"success": False, "error": result.get("error", "Query failed")}

            columns = result["columns"]
            rows = result["rows"]
            if not rows:
                return {"success": False, "error": "Query returned no data"}

            # Find the split column index
            split_col = split_by_column.strip()
            try:
                split_idx = next(i for i, c in enumerate(columns) if c.lower() == split_col.lower())
            except StopIteration:
                return {"success": False, "error": f"Column '{split_by_column}' not found. Available: {columns}"}

            # Remove split column from output columns
            out_columns = [c for i, c in enumerate(columns) if i != split_idx]

            # Group rows by split column value
            from collections import OrderedDict
            groups = OrderedDict()
            for row in rows:
                key = str(row[split_idx] or "Other")
                if key not in groups:
                    groups[key] = []
                groups[key].append([v for i, v in enumerate(row) if i != split_idx])

            for group_name, group_rows in groups.items():
                sheet_name = str(group_name)[:31]
                ws = wb.create_sheet(title=sheet_name)
                self._write_sheet(ws, out_columns, group_rows)
                total_rows += len(group_rows)
                sheet_results.append({"sheet": sheet_name, "rows": len(group_rows)})
        else:
            # Standard mode: one query per sheet
            for q in queries:
                sheet_name = str(q.get("name", "Sheet"))[:31]
                sql = q.get("sql", "")
                database = q.get("database")

                if not sql:
                    sheet_results.append({"sheet": sheet_name, "error": "No SQL provided"})
                    continue

                result = executor.execute(sql, database=database, _internal_limit=50000)
                if not result.get("success"):
                    sheet_results.append({"sheet": sheet_name, "error": result.get("error", "Query failed")})
                    continue

                columns = result["columns"]
                rows = result["rows"]
                if not rows:
                    sheet_results.append({"sheet": sheet_name, "rows": 0, "note": "No data returned"})
                    continue

                ws = wb.create_sheet(title=sheet_name)
                self._write_sheet(ws, columns, rows)
                total_rows += len(rows)
                sheet_results.append({"sheet": sheet_name, "rows": len(rows)})

        if total_rows == 0:
            return {"success": False, "error": "All queries returned no data", "details": sheet_results}

        filepath = self.export_path / xlsx_name
        wb.save(filepath)

        return {
            "success": True,
            "file_id": xlsx_name,
            "download_url": f"/download/{xlsx_name}",
            "row_count": total_rows,
            "sheet_count": len([s for s in sheet_results if s.get("rows", 0) > 0]),
            "sheets": sheet_results,
            "message": f"Excel file ready: {xlsx_name} ({total_rows} total rows)"
        }


# =============================================================================
# HARVEST PLANNER API CLIENT
# =============================================================================

class HarvestPlannerAPI:
    """HTTP client for the Harvest Planner REST API with JWT auth."""

    def __init__(self, base_url: str = CONFIG["hp_base_url"],
                 username: str = CONFIG["hp_username"],
                 password: str = CONFIG["hp_password"]):
        self.base_url = base_url
        self.username = username
        self.password = password
        self.access_token: Optional[str] = None
        self.refresh_token: Optional[str] = None
        self.token_expires_at: Optional[datetime] = None
        self.client = httpx.Client(timeout=30)

    def _is_configured(self) -> bool:
        return bool(self.base_url and self.username and self.password)

    def _ensure_auth(self) -> Optional[str]:
        """Login or refresh token as needed. Returns error string or None."""
        if not self._is_configured():
            return "Harvest Planner API not configured. Set HP_BASE_URL, HP_USERNAME, HP_PASSWORD in .env"

        # If we have a valid token, use it
        if self.access_token and self.token_expires_at:
            if datetime.now(timezone.utc) < self.token_expires_at:
                return None
            # Try refresh
            if self.refresh_token:
                err = self._refresh()
                if err is None:
                    return None

        # Login fresh
        return self._login()

    def _login(self) -> Optional[str]:
        try:
            resp = self.client.post(
                f"{self.base_url}/api/v1/auth/login",
                json={"username": self.username, "password": self.password}
            )
            if resp.status_code == 401:
                return "Harvest Planner login failed: bad credentials"
            if resp.status_code == 423:
                return "Harvest Planner account is locked"
            content_type = resp.headers.get("content-type", "")
            if "application/json" not in content_type:
                return f"Harvest Planner login returned non-JSON ({resp.status_code}): {resp.text[:200]}"
            resp.raise_for_status()
            data = resp.json()
            self.access_token = data["accessToken"]
            self.refresh_token = data["refreshToken"]
            self.token_expires_at = datetime.fromisoformat(
                data["accessTokenExpiresAt"].replace("Z", "+00:00")
            )
            return None
        except Exception as e:
            return f"Harvest Planner login error: {e}"

    def _refresh(self) -> Optional[str]:
        try:
            resp = self.client.post(
                f"{self.base_url}/api/v1/auth/refresh",
                json={"refreshToken": self.refresh_token}
            )
            if resp.status_code != 200:
                return f"Token refresh failed (HTTP {resp.status_code})"
            content_type = resp.headers.get("content-type", "")
            if "application/json" not in content_type:
                return f"Token refresh returned non-JSON: {resp.text[:200]}"
            data = resp.json()
            self.access_token = data["accessToken"]
            self.refresh_token = data["refreshToken"]
            self.token_expires_at = datetime.fromisoformat(
                data["accessTokenExpiresAt"].replace("Z", "+00:00")
            )
            return None
        except Exception as e:
            return f"Token refresh error: {e}"

    def _headers(self) -> dict:
        return {"Authorization": f"Bearer {self.access_token}"}

    def _request(self, method: str, path: str, params: dict = None,
                 json_body: dict = None, auth_required: bool = True) -> dict:
        """Make an API request with automatic auth handling."""
        if auth_required:
            err = self._ensure_auth()
            if err:
                return {"success": False, "error": err}

        url = f"{self.base_url}{path}"
        headers = self._headers() if auth_required and self.access_token else {}

        try:
            resp = self.client.request(method, url, params=params,
                                       json=json_body, headers=headers)
            if resp.status_code == 401 and auth_required:
                # Token may have expired mid-request, retry with fresh login
                err = self._login()
                if err:
                    return {"success": False, "error": err}
                headers = self._headers()
                resp = self.client.request(method, url, params=params,
                                           json=json_body, headers=headers)

            if resp.status_code == 204:
                return {"success": True, "message": "Operation completed successfully"}

            # Guard against non-JSON responses (HTML error pages, proxy errors)
            content_type = resp.headers.get("content-type", "")
            if "application/json" not in content_type:
                snippet = resp.text[:300].strip()
                return {"success": False,
                        "error": f"HTTP {resp.status_code} - expected JSON but got {content_type or 'unknown content type'}",
                        "detail": snippet}

            if resp.status_code >= 400:
                try:
                    body = resp.json()
                except Exception:
                    body = resp.text
                return {"success": False, "error": f"HTTP {resp.status_code}", "detail": body}

            return {"success": True, "data": resp.json()}
        except httpx.ConnectError as e:
            return {"success": False, "error": f"Connection failed to {self.base_url}: {e}"}
        except httpx.TimeoutException:
            return {"success": False, "error": f"Request timed out ({self.client.timeout}s) to {url}"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    # --- Harvest Plans (auth required) ---
    def create_harvest_plan(self, plan_data: dict) -> dict:
        return self._request("POST", "/api/v1/harvestplans", json_body=plan_data)

    def update_harvest_plan(self, plan_id: str, plan_data: dict) -> dict:
        return self._request("PUT", f"/api/v1/harvestplans/{plan_id}",
                             json_body=plan_data)

    def delete_harvest_plan(self, plan_id: str) -> dict:
        return self._request("DELETE", f"/api/v1/harvestplans/{plan_id}")

    # --- Harvest Contractors ---
    def create_contractor(self, contractor_data: dict) -> dict:
        return self._request("POST", "/api/v1/harvestcontractors",
                             json_body=contractor_data, auth_required=False)

    # --- Placeholder Growers (auth required) ---
    def create_placeholder_grower(self, grower_data: dict) -> dict:
        return self._request("POST", "/api/v1/placeholdergrower",
                             json_body=grower_data)

    # --- Production Runs (auth required) ---
    def create_production_run(self, run_data: dict) -> dict:
        return self._request("POST", "/api/v1/productionruns", json_body=run_data)



# =============================================================================
# LIVING DOCUMENT HELPERS (for agent tool use)
# =============================================================================

def _agent_update_living_doc(name: str, new_name: str = None,
                             new_description: str = None, new_prompt: str = None) -> dict:
    doc = _living_docs.get_doc_by_name(name)
    if not doc:
        available = [d["name"] for d in _living_docs.list_docs()]
        return {"error": f"No living document named '{name}'.", "available": available}
    updated = _living_docs.update_doc(doc["id"], name=new_name,
                                      description=new_description, prompt=new_prompt)
    if not updated:
        return {"error": "No fields to update."}
    return {"success": True, "document": updated}


def _agent_delete_living_doc(name: str) -> dict:
    doc = _living_docs.get_doc_by_name(name)
    if not doc:
        available = [d["name"] for d in _living_docs.list_docs()]
        return {"error": f"No living document named '{name}'.", "available": available}
    _living_docs.delete_doc(doc["id"])
    return {"success": True, "deleted": name}


def _agent_get_living_doc(name: str) -> dict:
    doc = _living_docs.get_doc_by_name(name)
    if not doc:
        available = [d["name"] for d in _living_docs.list_docs()]
        return {
            "error": f"No living document named '{name}'.",
            "available_documents": available,
        }
    snap = _living_docs.get_latest_snapshot(doc["id"])
    if not snap:
        return {
            "name": doc["name"],
            "description": doc.get("description"),
            "snapshot": None,
            "message": "No snapshot generated yet. Ask the user to click Refresh in the sidebar.",
        }
    return {
        "name": doc["name"],
        "description": doc.get("description"),
        "snapshot": snap,
    }


# =============================================================================
# AGENT TOOLKIT - Main Interface
# =============================================================================

class AgentToolkit:
    """
    Main interface for AI agent tools.
    Combines query execution, context loading, and learning.
    """
    
    def __init__(self):
        self.executor = QueryExecutor()
        self.context = ContextLoader()
        self.learning = LearningManager()
        self.exporter = ExcelExporter()
        self.harvest_planner = HarvestPlannerAPI()
    
    def get_tool_definitions(self) -> list[dict]:
        """Return tool definitions for LLM function calling."""
        return [
            # === QUERY TOOLS ===
            {
                "name": "execute_sql",
                "description": "Execute a read-only SQL query against a data warehouse. Only SELECT queries are allowed. Returns columns, rows, and row count. Defaults to DM03. Use database='DM01' for harvest planning data (dbo.harvestplanentry, dbo.harvestcontractors, dbo.processproductionruns).",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "sql": {
                            "type": "string",
                            "description": "The SQL SELECT query to execute"
                        },
                        "max_rows": {
                            "type": "integer",
                            "description": "Maximum rows to return (default 1000, max 5000)"
                        },
                        "database": {
                            "type": "string",
                            "enum": ["DM03", "DM01"],
                            "description": "Target database. DM03 (default) for inventory/sales/operations. DM01 for harvest planning data."
                        }
                    },
                    "required": ["sql"]
                }
            },
            
            # === CONTEXT TOOLS ===
            {
                "name": "get_system_architecture",
                "description": "Load the system architecture document. Contains critical info about CF/LP systems, join rules, and column naming conventions. READ THIS FIRST before writing queries.",
                "parameters": {
                    "type": "object",
                    "properties": {}
                }
            },
            {
                "name": "get_glossary",
                "description": "Load the business glossary with definitions for domain terms like Bin, Carton, Pool, Grower, Packout, etc.",
                "parameters": {
                    "type": "object",
                    "properties": {}
                }
            },
            {
                "name": "get_table_schema",
                "description": "Load detailed schema documentation for a specific table including column descriptions, relationships, and sample queries.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "table_name": {
                            "type": "string",
                            "description": "Table name (e.g., 'VW_BININVENTORY', 'BININVSNAPSHOTS_EXT')"
                        }
                    },
                    "required": ["table_name"]
                }
            },
            {
                "name": "search_context",
                "description": "Search across all documentation for a term or concept.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {
                            "type": "string",
                            "description": "Search term to find in documentation"
                        }
                    },
                    "required": ["query"]
                }
            },
            {
                "name": "list_available_tables",
                "description": "List all documented tables in the data warehouse.",
                "parameters": {
                    "type": "object",
                    "properties": {}
                }
            },
            {
                "name": "get_size_dictionary",
                "description": "Get valid fruit sizes by commodity from local reference data. Use this INSTEAD of querying the database when the user asks about sizes, size lists, or size ranges for a commodity. Returns all valid size codes.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "commodity": {
                            "type": "string",
                            "description": "Optional commodity name (e.g., 'NAVEL', 'MANDARIN'). Omit to get all commodities."
                        }
                    }
                }
            },

            # === LEARNING TOOLS ===
            {
                "name": "remember_query",
                "description": "Save a successful query pattern for future reference. Call this when a query works well so you can reuse the pattern.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "question": {
                            "type": "string",
                            "description": "The natural language question this query answers"
                        },
                        "sql": {
                            "type": "string",
                            "description": "The SQL query that worked"
                        },
                        "notes": {
                            "type": "string",
                            "description": "Any notes about why this works or gotchas"
                        }
                    },
                    "required": ["question", "sql"]
                }
            },
            {
                "name": "log_correction",
                "description": "Log when you were corrected about something. This helps improve documentation over time.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "table": {
                            "type": "string",
                            "description": "Table the correction applies to"
                        },
                        "column": {
                            "type": "string",
                            "description": "Column name if applicable"
                        },
                        "wrong_assumption": {
                            "type": "string",
                            "description": "What you incorrectly assumed"
                        },
                        "correct_meaning": {
                            "type": "string",
                            "description": "The correct interpretation"
                        }
                    },
                    "required": ["table", "wrong_assumption", "correct_meaning"]
                }
            },
            {
                "name": "record_discovery",
                "description": "Record a new discovery about the data, schema, or business rules.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "category": {
                            "type": "string",
                            "enum": ["column_meaning", "join_pattern", "gotcha", "business_rule"],
                            "description": "Category of discovery"
                        },
                        "key": {
                            "type": "string",
                            "description": "Short identifier (e.g., column name, table name)"
                        },
                        "value": {
                            "type": "string",
                            "description": "What you discovered"
                        }
                    },
                    "required": ["category", "key", "value"]
                }
            },
            {
                "name": "get_learned_queries",
                "description": "Retrieve previously successful query patterns. Check this before writing a new query.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "search": {
                            "type": "string",
                            "description": "Optional search term to filter queries"
                        }
                    }
                }
            },
            {
                "name": "get_discoveries",
                "description": "Retrieve recorded discoveries about the data.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "category": {
                            "type": "string",
                            "enum": ["column_meaning", "join_pattern", "gotcha", "business_rule"],
                            "description": "Filter by category"
                        }
                    }
                }
            },
            {
                "name": "save_customer_spec",
                "description": "Save a customer-specific specification or preference. Use when the user tells you about a customer's size requirements, grade tolerances, packaging preferences, DC-specific rules, or any other customer-specific rule. Examples: 'Costco Mira Loma can take 88s', 'Safeway Portland is tougher on grade'.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "customer": {
                            "type": "string",
                            "description": "Customer name (e.g., 'Costco', 'Safeway', 'Walmart')"
                        },
                        "spec_type": {
                            "type": "string",
                            "enum": ["size", "grade", "packaging", "label", "pallet", "general"],
                            "description": "Type of specification"
                        },
                        "rule": {
                            "type": "string",
                            "description": "The specific rule or preference in plain language"
                        },
                        "dc": {
                            "type": "string",
                            "description": "Distribution center or location if the rule is DC-specific (e.g., 'Mira Loma', 'Sumner', 'Portland')"
                        }
                    },
                    "required": ["customer", "spec_type", "rule"]
                }
            },
            {
                "name": "get_customer_specs",
                "description": "Retrieve saved customer specifications and preferences. Check this when building production schedules, making size substitution recommendations, or answering questions about what a customer accepts.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "customer": {
                            "type": "string",
                            "description": "Optional customer name to filter by"
                        }
                    }
                }
            },

            # === EXPORT TOOLS ===
            {
                "name": "export_excel",
                "description": "Export data to a downloadable Excel (.xlsx) file. Supports single-sheet (columns/rows) or multi-sheet (sheets array) exports. For multi-sheet, pass a 'sheets' array where each element has name, columns, and rows. Always use multi-sheet when the user asks for multiple tables or breakdowns in one file. Returns a download link to include in your response.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "columns": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Column headers (single-sheet mode). Omit if using 'sheets'."
                        },
                        "rows": {
                            "type": "array",
                            "items": {"type": "array"},
                            "description": "Data rows (single-sheet mode). Omit if using 'sheets'."
                        },
                        "sheets": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "name": {"type": "string", "description": "Sheet/tab name (max 31 chars)"},
                                    "columns": {"type": "array", "items": {"type": "string"}},
                                    "rows": {"type": "array", "items": {"type": "array"}}
                                },
                                "required": ["name", "columns", "rows"]
                            },
                            "description": "Array of sheets for multi-sheet export. Each sheet has a name, columns, and rows."
                        },
                        "filename": {
                            "type": "string",
                            "description": "Descriptive filename (without extension), e.g. 'navel_inventory_march'"
                        }
                    }
                }
            },

            {
                "name": "export_sql_to_excel",
                "description": "Run SQL queries and export results DIRECTLY to Excel — data goes straight from database to spreadsheet. Use this instead of export_excel for complete exports. Supports two modes: (1) Multiple queries, each becoming a sheet. (2) Single query with split_by_column to auto-split results into one sheet per distinct value of that column (e.g., split by Style to get one tab per style). Use split_by_column for production schedules.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "queries": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "name": {"type": "string", "description": "Sheet/tab name (max 31 chars). Ignored when split_by_column is used."},
                                    "sql": {"type": "string", "description": "SELECT query to run"},
                                    "database": {"type": "string", "description": "Database to query (DM03 or DM01). Defaults to DM03."}
                                },
                                "required": ["name", "sql"]
                            },
                            "description": "Array of queries. When split_by_column is set, only the first query is used."
                        },
                        "split_by_column": {
                            "type": "string",
                            "description": "Column name to split results by. Each distinct value becomes a separate sheet tab. The column is removed from the sheet data. E.g., 'Style' splits orders into one sheet per style. For production schedules, include a Style column in the query and set this to 'Style'."
                        },
                        "filename": {
                            "type": "string",
                            "description": "Descriptive filename (without extension)"
                        }
                    },
                    "required": ["queries"]
                }
            },

            # === HARVEST PLANNER WRITE TOOLS (via API) ===
            {
                "name": "hp_create_harvest_plan",
                "description": "Create a new harvest plan via the Harvest Planner API. Link a grower block (or placeholder grower) with contractors, rates, pool, and scheduling. Fields: grower_block_source_database, grower_block_id (GABLOCKIDX), placeholder_grower_id (GUID, use instead of block if grower not in system), field_representative_id (user ID), planned_bins, contractor_id, harvesting_rate, hauler_id, hauling_rate, forklift_contractor_id, forklift_rate, pool_id (POOLIDX), notes_general, deliver_to, packed_by, date (YYYY-MM-DD), bins.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "plan_data": {
                            "type": "object",
                            "description": "Harvest plan fields to set"
                        }
                    },
                    "required": ["plan_data"]
                }
            },
            {
                "name": "hp_update_harvest_plan",
                "description": "Update an existing harvest plan via the Harvest Planner API. Pass only the fields to change.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "plan_id": {"type": "string", "description": "The harvest plan GUID to update"},
                        "plan_data": {
                            "type": "object",
                            "description": "Fields to update"
                        }
                    },
                    "required": ["plan_id", "plan_data"]
                }
            },
            {
                "name": "hp_delete_harvest_plan",
                "description": "Delete a harvest plan via the Harvest Planner API by its ID (GUID).",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "plan_id": {"type": "string", "description": "The harvest plan GUID to delete"}
                    },
                    "required": ["plan_id"]
                }
            },
            {
                "name": "hp_create_contractor",
                "description": "Create a new harvest contractor via the Harvest Planner API. Fields: name (required), primary_contact_name, primary_contact_phone, office_phone, mailing_address, provides_trucking (bool), provides_picking (bool), provides_forklift (bool).",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "contractor_data": {
                            "type": "object",
                            "description": "Contractor fields to set"
                        }
                    },
                    "required": ["contractor_data"]
                }
            },
            {
                "name": "hp_create_placeholder_grower",
                "description": "Create a placeholder grower via the Harvest Planner API for use in harvest plans when the real block doesn't exist yet. Fields: grower_name (required), commodity_name (required), is_active (default true), notes.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "grower_data": {
                            "type": "object",
                            "description": "Placeholder grower fields"
                        }
                    },
                    "required": ["grower_data"]
                }
            },
            {
                "name": "hp_create_production_run",
                "description": "Create a production run via the Harvest Planner API to track processing/packing of harvested fruit. Fields: source_database (required), gablockidx (required, > 0), bins, run_date, pick_date, location, pool, notes, row_order, run_status, batch_id, time_started, time_completed.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "run_data": {
                            "type": "object",
                            "description": "Production run fields"
                        }
                    },
                    "required": ["run_data"]
                }
            },

            # === LIVING DOCUMENT TOOLS ===
            {
                "name": "list_living_documents",
                "description": "List all defined living documents (shared daily reports consistent across all users). Use this to show the user what documents exist or to check before creating a duplicate.",
                "parameters": {
                    "type": "object",
                    "properties": {}
                }
            },
            {
                "name": "get_living_document",
                "description": "Retrieve today's snapshot of a living document by name. Living documents are shared across all users and refreshed daily. Use this when a user asks to see a living document or wants to discuss its contents.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {
                            "type": "string",
                            "description": "Exact name of the living document (use list_living_documents if unsure)"
                        }
                    },
                    "required": ["name"]
                }
            },
            {
                "name": "create_living_document",
                "description": "Register a new living document definition. Called when the user types /living-doc-add. The prompt must be a complete, self-contained instruction that will be run daily to generate the document (e.g., 'Generate a daily production summary showing all packing lines...'). Confirm the name and prompt with the user before calling this.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {
                            "type": "string",
                            "description": "Short display name (e.g., 'Daily Production Plan', 'Morning Pick Plan')"
                        },
                        "description": {
                            "type": "string",
                            "description": "One-sentence description of what this document shows"
                        },
                        "prompt": {
                            "type": "string",
                            "description": "The full prompt to run to generate this document. Must be detailed and self-contained."
                        }
                    },
                    "required": ["name", "prompt"]
                }
            },
            {
                "name": "update_living_document",
                "description": "Update an existing living document's name, description, or prompt. Use when the user wants to change what a living document generates or rename it. Look up the document by name first using get_living_document or list_living_documents to get the ID.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {
                            "type": "string",
                            "description": "Current name of the living document to update"
                        },
                        "new_name": {
                            "type": "string",
                            "description": "New display name (omit to keep current)"
                        },
                        "new_description": {
                            "type": "string",
                            "description": "New description (omit to keep current)"
                        },
                        "new_prompt": {
                            "type": "string",
                            "description": "New prompt to generate the document (omit to keep current)"
                        }
                    },
                    "required": ["name"]
                }
            },
            {
                "name": "delete_living_document",
                "description": "Delete a living document by name. Confirm with the user before calling this. The document and its snapshots will no longer be visible.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {
                            "type": "string",
                            "description": "Name of the living document to delete"
                        }
                    },
                    "required": ["name"]
                }
            },
        ]
    
    def handle_tool_call(self, tool_name: str, parameters: dict) -> dict:
        """Route a tool call to the appropriate handler."""
        
        handlers = {
            # Query tools
            "execute_sql": lambda p: self.executor.execute(
                p.get("sql", ""), p.get("max_rows"), p.get("database")
            ),
            
            # Context tools
            "get_system_architecture": lambda p: self.context.load_system_architecture(),
            "get_glossary": lambda p: self.context.load_glossary(),
            "get_size_dictionary": lambda p: self.context.load_size_dictionary(p.get("commodity")),
            "get_table_schema": lambda p: self.context.load_table_schema(p.get("table_name", "")),
            "search_context": lambda p: self.context.search_context(p.get("query", "")),
            "list_available_tables": lambda p: self.context.list_tables(),
            
            # Learning tools
            "remember_query": lambda p: self.learning.remember_query(
                p.get("question", ""), p.get("sql", ""), p.get("notes", "")
            ),
            "log_correction": lambda p: self.learning.log_correction(
                p.get("table", ""), p.get("column", ""),
                p.get("wrong_assumption", ""), p.get("correct_meaning", "")
            ),
            "record_discovery": lambda p: self.learning.record_discovery(
                p.get("category", ""), p.get("key", ""), p.get("value", "")
            ),
            "get_learned_queries": lambda p: self.learning.get_learned_queries(p.get("search")),
            "get_discoveries": lambda p: self.learning.get_discoveries(p.get("category")),
            "save_customer_spec": lambda p: _customer_specs.save_spec(
                p.get("customer", ""), p.get("spec_type", "general"),
                p.get("rule", ""), p.get("dc")
            ),
            "get_customer_specs": lambda p: _customer_specs.get_specs(p.get("customer")),

            # Export tools
            "export_excel": lambda p: self.exporter.export(
                p.get("columns"), p.get("rows"), p.get("filename", ""), p.get("sheets")
            ),
            "export_sql_to_excel": lambda p: self.exporter.export_from_queries(
                p.get("queries", []), p.get("filename", ""), p.get("split_by_column")
            ),

            # Harvest Planner write tools (via API)
            "hp_create_harvest_plan": lambda p: self.harvest_planner.create_harvest_plan(
                p.get("plan_data", {})
            ),
            "hp_update_harvest_plan": lambda p: self.harvest_planner.update_harvest_plan(
                p.get("plan_id", ""), p.get("plan_data", {})
            ),
            "hp_delete_harvest_plan": lambda p: self.harvest_planner.delete_harvest_plan(
                p.get("plan_id", "")
            ),
            "hp_create_contractor": lambda p: self.harvest_planner.create_contractor(
                p.get("contractor_data", {})
            ),
            "hp_create_placeholder_grower": lambda p: self.harvest_planner.create_placeholder_grower(
                p.get("grower_data", {})
            ),
            "hp_create_production_run": lambda p: self.harvest_planner.create_production_run(
                p.get("run_data", {})
            ),

            # Living document tools
            "list_living_documents": lambda p: _living_docs.list_docs(),
            "get_living_document": lambda p: _agent_get_living_doc(p.get("name", "")),
            "create_living_document": lambda p: _living_docs.create_doc(
                p.get("name", ""), p.get("description", ""), p.get("prompt", ""),
                created_by="agent",
            ),
            "update_living_document": lambda p: _agent_update_living_doc(
                p.get("name", ""), p.get("new_name"), p.get("new_description"), p.get("new_prompt"),
            ),
            "delete_living_document": lambda p: _agent_delete_living_doc(p.get("name", "")),
        }
        
        handler = handlers.get(tool_name)
        if not handler:
            return {"error": f"Unknown tool: {tool_name}"}
        
        return handler(parameters)


# =============================================================================
# CLI FOR TESTING
# =============================================================================

if __name__ == "__main__":
    toolkit = AgentToolkit()
    
    print("DM03 Agent Toolkit - Interactive Test Mode")
    print("=" * 50)
    print("\nAvailable tools:")
    for tool in toolkit.get_tool_definitions():
        print(f"  - {tool['name']}: {tool['description'][:60]}...")
    
    print("\n\nTest: Loading system architecture...")
    result = toolkit.handle_tool_call("get_system_architecture", {})
    if "error" not in result:
        print("  ✓ System architecture loaded")
    else:
        print(f"  ✗ Error: {result['error']}")
    
    print("\nTest: Loading glossary...")
    result = toolkit.handle_tool_call("get_glossary", {})
    if "error" not in result:
        print("  ✓ Glossary loaded")
    else:
        print(f"  ✗ Error: {result['error']}")
    
    print("\nTest: Recording a discovery...")
    result = toolkit.handle_tool_call("record_discovery", {
        "category": "gotcha",
        "key": "test_key",
        "value": "This is a test discovery"
    })
    print(f"  {result}")
    
    print("\nToolkit ready for agent integration.")